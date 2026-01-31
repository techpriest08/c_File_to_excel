import openpyxl
import os

# 1. C 및 H 파일들이 들어있는 '폴더' 경로를 직접 지정하세요
folder_path = r"파일 위치"
output_file = "저장될 엑셀 이름"

def run():
    try:
        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()
        
        # 2. 폴더 내의 모든 .c 및 .h 파일 목록 가져오기
        # endswith에 튜플 ('.c', '.h')를 전달하여 두 확장자를 모두 찾습니다.
        file_list = [f for f in os.listdir(folder_path) if f.lower().endswith(('.c', '.h'))]
        
        if not file_list:
            print(f"지정한 경로에 .c 또는 .h 파일이 없습니다.\n경로: {folder_path}")
            return

        # 파일 이름순으로 정렬
        file_list.sort()

        for file_name in file_list:
            print(f"처리 중: {file_name}")
            
            # 시트 이름 설정 (최대 31자 제한)
            ws = wb.create_sheet(title=file_name[:31])
            
            file_full_path = os.path.join(folder_path, file_name)
            
            # 파일 읽기 (cp949 시도 후 안되면 utf-8)
            try:
                with open(file_full_path, "r", encoding="cp949") as f:
                    lines = f.readlines()
            except:
                with open(file_full_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()

            # 시트에 내용 채우기
            for i, line in enumerate(lines, start=1):
                ws.cell(row=i, column=1, value=line.rstrip("\n\r"))
            
            # 열 너비 조절
            ws.column_dimensions['A'].width = 100

        # 처음에 자동으로 만들어진 빈 시트 삭제
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 3. 엑셀 파일 저장 (파이썬 파일 위치에 저장됨)
        wb.save(output_file)
        print("-" * 40)
        print(f"모든 작업 완료! 생성된 파일: {output_file}")
        print(f"총 {len(file_list)}개의 파일이 변환되었습니다.")

    except Exception as e:
        print(f"오류 발생: {e}")

    finally:
        # 이 부분이 추가되었습니다. 어떤 상황에서도 창이 바로 닫히지 않게 잡아줍니다.
        print("\n" + "="*40)
        input("창을 닫으려면 [Enter] 키를 누르세요...")

if __name__ == "__main__":
    run()